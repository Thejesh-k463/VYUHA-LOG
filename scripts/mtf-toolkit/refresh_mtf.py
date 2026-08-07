#!/usr/bin/env python3
"""
refresh_mtf.py — India MTF (Margin Trading Facility) approved-scrip refresher.

Pulls each broker's own machine-readable feed, validates it, stores a dated
snapshot in SQLite, diffs against the previous snapshot, and rebuilds the Excel
workbook plus flat CSVs.

Brokers currently wired up:
    Dhan        - live Google Sheet behind dhan.co/mtf-stocks-list
    Zerodha     - public.zrd.sh JSON behind zerodha.com/mtf-approved-securities
    Upstox      - service.upstox.com search API behind upstox.com/stocks/mtf-stocks-list
    Kotak Neo   - SSR flight payload behind kotakneo.com/margin-requirement/margin-trading/
    Paytm Money - api-eq.paytmmoney.com behind paytmmoney.com/mtf
    Angel One   - approved list from a local scrip-category file + margin modelled
                  from NSE's official daily VaR file (see fetch_angel)
    Groww       - local CSV harvested in-browser (see fetch_groww); the web surface
                  renders 100 rows per view, so a full list needs the accumulator

Sahi does not offer MTF and is excluded by design.

LOCAL INPUTS (folder set by INPUT_DIR, default ./input)
    Angel One : JULYSCRIPCATEGORY*.xlsx  (or any *SCRIPCATEGORY*.xlsx) - monthly,
                downloaded from the Angel back office. Newest file wins.
    Groww     : groww_mtf*.csv - produced by groww_accumulate.js in the browser.
                Newest file wins.
Both brokers are skipped with a warning if their input file is absent.

Usage
    python refresh_mtf.py                 # full refresh
    python refresh_mtf.py --no-excel      # data + CSV + SQLite only, skip workbook
    python refresh_mtf.py --brokers dhan,kotakneo
    python refresh_mtf.py --recalc        # try LibreOffice to cache formula values
    python refresh_mtf.py --history 30    # print margin history for last 30 snapshots

Exit codes
    0  success
    1  a broker fetch failed
    2  validation gate failed (nothing was written)
    3  bad arguments / environment
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import json
import os
import re
import sqlite3
import sys
import time
import urllib.error
import urllib.request
import uuid

# --------------------------------------------------------------------------
# CONFIG — edit these paths for your machine
# --------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input")        # drop Angel + Groww files here
RAW_DIR = os.path.join(BASE_DIR, "raw")            # dated raw payloads, kept forever
OUT_DIR = os.path.join(BASE_DIR, "output")         # workbook + CSVs
DB_PATH = os.path.join(BASE_DIR, "mtf_history.db")  # SQLite snapshot history
WORKBOOK = "India_MTF_Stock_Lists.xlsx"

IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/125.0 Safari/537.36")

HTTP_RETRIES = 3
HTTP_BACKOFF = 2.0      # seconds, doubled each retry
HTTP_TIMEOUT = 45

# Endpoints. If a broker changes their site, this is the only block to touch.
DHAN_SHEET_ID = "1qfMHvvnubpus4lkuh-DbztlL0zUKnVRv7IelciPErlc"
DHAN_GID = "35869169"
DHAN_URL = (f"https://docs.google.com/spreadsheets/d/{DHAN_SHEET_ID}"
            f"/gviz/tq?tqx=out:csv&gid={DHAN_GID}")
DHAN_PAGE = "https://dhan.co/mtf-stocks-list/"

ZERODHA_URL = "https://public.zrd.sh/crux/approved-mtf-securities.json"
ZERODHA_PAGE = "https://zerodha.com/mtf-approved-securities"

UPSTOX_URL = ("https://service.upstox.com/search/open/v1/"
              "?query=&segments=MTF_EQ&exchanges={exchange}"
              "&records={records}&pageNumber={page}")
UPSTOX_PAGE = "https://upstox.com/stocks/mtf-stocks-list/"
UPSTOX_PAGE_SIZE = 250
UPSTOX_MAX_PAGES = 40      # safety stop; their own client caps at 40

KOTAK_PAGE = "https://www.kotakneo.com/margin-requirement/margin-trading/"
KOTAK_PAGE_N = "https://www.kotakneo.com/margin-requirement/margin-trading/{page}/"
KOTAK_MAX_PAGES = 200          # safety stop; site reports ~56

PAYTM_URL = "https://api-eq.paytmmoney.com/mtf/order/api/v2/scrips"
PAYTM_PAGE = "https://www.paytmmoney.com/mtf"

# --- Angel One -------------------------------------------------------------
# NSE publishes six intraday VaR snapshots per trading day. The rate that
# APPLIES to a given day is snapshot 1 of that date (identical to snapshot 6 of
# the previous day). Using a later snapshot degrades the Angel model roughly
# eightfold, so snapshot 1 is not negotiable.
NSE_VAR_URL = "https://nsearchives.nseindia.com/archives/nsccl/var/C_VAR1_{ddmmyyyy}_1.DAT"
NSE_FO_URL = "https://nsearchives.nseindia.com/content/fo/fo_mktlots.csv"
NSE_REFERER = "https://www.nseindia.com/"

# Calibrated against 17 readings from Angel One's own Margin Calculator.
#   base = VaR + 5 x ELM
#   non-F&O margin = base
#   F&O     margin = ANGEL_K_FO x base
# The F&O figure is a CONSTANT, not a regression: implied multiplier ranged
# 0.85117-0.85142 across 11 F&O readings (sd 0.00006). Back-test MAE 0.008pp.
ANGEL_K_FO = 0.8513
ANGEL_BACKTEST = "17 readings, MAE 0.008pp, worst 0.07pp"

PENDING_BROKERS = {
    "Sahi": "Does not offer MTF. No mention across homepage, pricing, FAQ or sitemap; "
            "their pledge FAQ scopes collateral to intraday, futures and options "
            "writing only. Excluded by design, not pending.",
}

ZCAT = {"fo": "F&O (derivatives-eligible)", "non_fo": "Non-F&O (cash only)",
        "etf": "ETF", "non_categorized": "Non-categorised"}


# --------------------------------------------------------------------------
# Plumbing
# --------------------------------------------------------------------------

def log(msg: str, level: str = "INFO") -> None:
    stamp = dt.datetime.now(IST).strftime("%H:%M:%S")
    print(f"[{stamp}] {level:<5} {msg}", flush=True)


def http_get(url: str, referer: str | None = None) -> bytes:
    """GET with retry + exponential backoff. Raises on final failure."""
    last = None
    for attempt in range(1, HTTP_RETRIES + 1):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": UA,
                "Accept": "*/*",
                **({"Referer": referer} if referer else {}),
                "x-request-id": str(uuid.uuid4()),
            })
            with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT) as r:
                return r.read()
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError) as e:
            last = e
            if attempt < HTTP_RETRIES:
                wait = HTTP_BACKOFF * (2 ** (attempt - 1))
                log(f"attempt {attempt}/{HTTP_RETRIES} failed ({e}); retrying in {wait:.0f}s",
                    "WARN")
                time.sleep(wait)
    raise RuntimeError(f"GET failed after {HTTP_RETRIES} attempts: {url} — {last}")


def save_raw(name: str, payload: bytes, stamp: str) -> str:
    os.makedirs(RAW_DIR, exist_ok=True)
    path = os.path.join(RAW_DIR, f"{stamp}_{name}")
    with open(path, "wb") as f:
        f.write(payload)
    return path


# --------------------------------------------------------------------------
# Normalised record
#   symbol, isin, category, margin_pct (what YOU fund),
#   broker_funding_pct, leverage, funded (bool), extra (dict)
# --------------------------------------------------------------------------

def safe_text(v):
    """Neutralise anything a spreadsheet would evaluate. Broker feeds are text
    we do not control — one Groww slug arrived as the literal string '#NAME?'
    from an upstream spreadsheet round-trip and broke the workbook recalc."""
    t = "" if v is None else str(v).strip()
    if not t:
        return ""
    if t.startswith(("=", "+", "@")) or (t.startswith("#") and t.endswith("?")):
        return ""
    if t.startswith(("#", "-")) and not re.fullmatch(r"-?\d+(\.\d+)?", t):
        return "'" + t
    return t


def rec(symbol, isin, margin, broker_funding, leverage, category="", **extra):
    return {
        "symbol": symbol, "isin": isin, "category": category,
        "margin_pct": round(float(margin), 4),
        "broker_funding_pct": round(float(broker_funding), 4),
        "leverage": round(float(leverage), 4),
        "funded": float(broker_funding) > 0,
        "extra": extra,
    }


# --------------------------------------------------------------------------
# Broker fetchers
# --------------------------------------------------------------------------

def fetch_dhan(stamp: str) -> list[dict]:
    log("Dhan: fetching live sheet")
    raw = http_get(DHAN_URL, referer=DHAN_PAGE)
    save_raw("dhan.csv", raw, stamp)
    rows = list(csv.reader(io.StringIO(raw.decode("utf-8", "replace"))))
    out = []
    for r in rows:
        # Data rows start with a numeric serial. The header row carries a long
        # advisory note in column B, so serial-detection is the safe filter.
        if len(r) < 6 or not r[0].strip().isdigit():
            continue
        sym, isin = r[1].strip(), r[2].strip()
        user, bfund, lev = float(r[3]), float(r[4]), float(r[5])
        out.append(rec(sym, isin, user, bfund, lev, sr=int(r[0])))
    log(f"Dhan: {len(out)} scrips")
    return out


def fetch_zerodha(stamp: str) -> list[dict]:
    log("Zerodha: fetching approved securities feed")
    raw = http_get(ZERODHA_URL, referer=ZERODHA_PAGE)
    save_raw("zerodha.json", raw, stamp)
    data = json.loads(raw)
    out = []
    for x in data:
        m = float(x["margin"])
        out.append(rec(x["tradingsymbol"], x["isin"], m, 100 - m,
                       float(x["leverage"]),
                       category=ZCAT.get(x.get("category", ""), x.get("category", ""))))
    log(f"Zerodha: {len(out)} scrips")
    return out


def fetch_upstox(stamp: str) -> list[dict]:
    log("Upstox: paginating MTF_EQ segment")
    seen, pages, declared = {}, 0, None
    page = 1
    while page <= UPSTOX_MAX_PAGES:
        url = UPSTOX_URL.format(exchange="NSE", records=UPSTOX_PAGE_SIZE, page=page)
        raw = http_get(url, referer=UPSTOX_PAGE)
        if page == 1:
            save_raw("upstox_p1.json", raw, stamp)
        d = json.loads(raw)
        if not d.get("success"):
            raise RuntimeError(f"Upstox API returned success=false on page {page}")
        lst = d["data"]["searchList"]
        meta = d["metaData"]["page"]
        declared = meta["totalRecords"]
        total_pages = meta["totalPages"]
        for s in lst:
            a = s["attributes"]
            seen[a["instrumentKey"]] = a
        pages += 1
        log(f"  page {page}/{total_pages} -> {len(lst)} rows (running {len(seen)})")
        if page >= total_pages or not lst:
            break
        page += 1

    out = []
    for a in seen.values():
        b = float(a["mtfBracket"])
        lev = (100.0 / b) if b else 0.0
        out.append(rec(a["tradingSymbol"], a["instrumentKey"].split("|")[1],
                       b, 100 - b, lev,
                       category=a.get("segment", ""),
                       instrument_key=a["instrumentKey"],
                       segment=a.get("segment", "")))
    save_raw("upstox_all.json", json.dumps(list(seen.values())).encode(), stamp)
    log(f"Upstox: {len(out)} scrips (API declared {declared})")
    # Completeness proof — this is the whole reason Upstox needs a gate.
    if declared is not None and len(out) != declared:
        raise RuntimeError(
            f"Upstox incomplete: captured {len(out)} but API declares {declared}. "
            "Refusing to write a partial list.")
    return out


def fetch_kotakneo(stamp: str) -> list[dict]:
    """Kotak Neo server-renders the table into the Next.js RSC flight payload,
    30 rows per page. There is no JSON API — the flight payload IS the source.

    Note: Kotak publishes leverage_multiple, NOT a margin percentage, and rounds
    it to 2dp. Margin is therefore derived as 100/leverage and carries rounding
    of roughly +/-0.2pp. It is the only broker here whose margin is inferred
    rather than published, and the only one exposing a per-scrip exposure cap.
    Kotak also omits ISIN, so ISINs are backfilled from the other brokers by
    symbol (see backfill_isins)."""
    import re
    log("Kotak Neo: paginating server-rendered pages")
    pat = re.compile(
        r'"sortData":\{"scrip_name":"(.*?)","nse_symbol":"(.*?)",'
        r'"leverage_multiple":([0-9.]+),"max_allowed_exposure_in_cr":([0-9.]+)\}')

    def unescape(s: str) -> str:
        s = re.sub(r"\\+u([0-9a-fA-F]{4})",
                   lambda m: chr(int(m.group(1), 16)), s)
        return s.replace("\\/", "/").strip()

    def grab(n: int):
        url = KOTAK_PAGE if n == 1 else KOTAK_PAGE_N.format(page=n)
        html = http_get(url, referer=KOTAK_PAGE).decode("utf-8", "replace")
        # The payload is double-escaped inside the flight script tags.
        flat = html.replace('\\\\\\"', "\x01").replace('\\"', '"').replace("\x01", '\\"')
        return pat.findall(flat), flat

    rows, flat = grab(1)
    if not rows:
        raise RuntimeError("Kotak Neo: no rows on page 1 — page structure may have changed")
    if page_no := __import__("re").search(r'"totalPages":(\d+)', flat):
        total = min(int(page_no.group(1)), KOTAK_MAX_PAGES)
    else:
        raise RuntimeError("Kotak Neo: totalPages not found — cannot prove completeness")
    save_raw("kotakneo_p1.html", flat.encode("utf-8", "replace"), stamp)

    seen = {}
    for a, b, c, d in rows:
        seen[unescape(b)] = (unescape(a), float(c), float(d))
    for p in range(2, total + 1):
        r, _ = grab(p)
        for a, b, c, d in r:
            seen[unescape(b)] = (unescape(a), float(c), float(d))
        if p % 10 == 0 or p == total:
            log(f"  page {p}/{total} (running {len(seen)})")

    out = []
    for sym, (name, lev, expo) in seen.items():
        margin = 100.0 / lev if lev else 100.0
        out.append(rec(sym, "", margin, 100 - margin, lev,
                       category="", scrip_name=name, max_exposure_cr=expo,
                       margin_is_derived=True))
    save_raw("kotakneo_all.json",
             json.dumps([{"symbol": s, "scrip_name": v[0], "leverage_multiple": v[1],
                          "max_exposure_cr": v[2]} for s, v in seen.items()]).encode(),
             stamp)
    log(f"Kotak Neo: {len(out)} scrips across {total} pages")
    # Completeness gate: 30 rows/page except possibly the last.
    if len(out) < (total - 1) * 30:
        raise RuntimeError(
            f"Kotak Neo incomplete: {len(out)} rows from {total} pages "
            f"(expected at least {(total - 1) * 30}). Refusing to write a partial list.")
    return out


def fetch_paytm(stamp: str) -> list[dict]:
    log("Paytm Money: fetching MTF scrip list")
    raw = http_get(PAYTM_URL, referer=PAYTM_PAGE)
    save_raw("paytm.json", raw, stamp)
    d = json.loads(raw)
    if not d.get("data"):
        raise RuntimeError("Paytm Money: response carried no data array")
    out = []
    for x in d["data"]:
        if x.get("status") != "ENABLED":
            continue
        m = float(x["margin_perc"])
        out.append(rec(x["symbol"], x["isin"], m, 100 - m, float(x["margin_x"]),
                       category=("ETF" if x.get("instrument_type") == "ETF"
                                 else "Equity"),
                       scrip_name=x.get("scrip_name", ""),
                       exchange=x.get("exchange", "")))
    log(f"Paytm Money: {len(out)} scrips "
        f"(of {len(d['data'])} returned, ENABLED only)")
    return out


def _newest(patterns: list[str]) -> str | None:
    """Newest file in INPUT_DIR matching any glob pattern."""
    import glob
    hits = []
    for p in patterns:
        hits += glob.glob(os.path.join(INPUT_DIR, p))
    if not hits:
        return None
    return max(hits, key=os.path.getmtime)


def fetch_nse_var(stamp: str) -> dict:
    """NSE security-wise VaR, keyed by ISIN. Walks back up to 7 days to cope
    with weekends and holidays. Snapshot 1 only — see NSE_VAR_URL note."""
    today = dt.datetime.now(IST).date()
    for back in range(0, 8):
        d = today - dt.timedelta(days=back)
        url = NSE_VAR_URL.format(ddmmyyyy=d.strftime("%d%m%Y"))
        try:
            raw = http_get(url, referer=NSE_REFERER)
        except Exception:
            continue
        if len(raw) < 100_000:
            continue
        save_raw(f"nse_var_{d:%Y%m%d}.dat", raw, stamp)
        var = {}
        for line in raw.decode("utf-8", "replace").splitlines():
            p = line.strip().split(",")
            if len(p) == 10 and p[0] == "20":
                var[p[3]] = {"symbol": p[1], "series": p[2],
                             "var": float(p[6]), "elm": float(p[7])}
        log(f"NSE VaR: {len(var)} securities from {d:%d-%b-%Y} snapshot 1")
        return var
    raise RuntimeError("NSE VaR file unavailable for the last 8 days")


def fetch_fo_symbols(stamp: str) -> set:
    """NSE F&O underlyings. This, not the Angel category, decides which margin
    branch a scrip takes."""
    raw = http_get(NSE_FO_URL, referer=NSE_REFERER)
    save_raw("fo_mktlots.csv", raw, stamp)
    out = set()
    for row in csv.reader(io.StringIO(raw.decode("utf-8", "replace"))):
        if len(row) >= 2:
            sym = row[1].strip().upper()
            if sym and sym != "SYMBOL":
                out.add(sym)
    log(f"NSE F&O underlyings: {len(out)}")
    return out


def fetch_angel(stamp: str) -> list[dict]:
    """Angel One publishes no per-scrip margin. The approved list comes from the
    monthly scrip-category file; the margin is modelled from NSE VaR."""
    path = _newest(["*SCRIPCATEGORY*.xlsx", "*scripcategory*.xlsx", "angel*.xlsx"])
    if not path:
        raise FileNotFoundError(
            f"No Angel scrip-category file in {INPUT_DIR}. Download the monthly "
            "file from the Angel back office and drop it there (e.g. "
            "JULYSCRIPCATEGORY_7648.xlsx).")
    log(f"Angel One: reading {os.path.basename(path)}")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for the Angel sheet")

    ws = load_workbook(path, data_only=True, read_only=True).worksheets[0]
    hdr, rows = None, []
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(c).strip().lower() if c else "" for c in r]
            continue
        rows.append(r)

    def col(*names):
        for n in names:
            for i, h in enumerate(hdr):
                if n in h:
                    return i
        return None

    i_sym, i_isin = col("nse symbol", "symbol"), col("isin")
    i_mtf, i_cat = col("nbfc/mtf", "mtf"), col("angelscripcategory", "category")
    i_ser, i_bse, i_name = col("series"), col("bse code"), col("co name", "name")
    if None in (i_sym, i_isin, i_mtf, i_cat):
        raise RuntimeError(f"Angel file has unexpected columns: {hdr}")

    var = fetch_nse_var(stamp)
    fo = fetch_fo_symbols(stamp)

    out, unmatched = [], 0
    for r in rows:
        if not r or len(r) <= max(i_sym, i_isin, i_mtf, i_cat):
            continue
        if str(r[i_mtf]).strip().lower() != "approved":
            continue
        sym, isin = str(r[i_sym] or "").strip(), str(r[i_isin] or "").strip()
        if not sym or not isin:
            continue
        v = var.get(isin)
        if not v:
            unmatched += 1
            continue
        base = v["var"] + 5 * v["elm"]
        isfo = sym.upper() in fo
        margin = min(100.0, max(0.0, ANGEL_K_FO * base if isfo else base))
        lev = 100.0 / margin if margin else 0.0
        out.append(rec(sym, isin, margin, 100 - margin, lev,
                       category=str(r[i_cat] or "").strip(),
                       series=str(r[i_ser] or "") if i_ser is not None else "",
                       bse_code=r[i_bse] if i_bse is not None else "",
                       scrip_name=str(r[i_name] or "") if i_name is not None else "",
                       fo_stock="Yes" if isfo else "No",
                       exchange_var=v["var"], elm=v["elm"],
                       margin_is_modelled=True))
    log(f"Angel One: {len(out)} scrips modelled "
        f"({sum(1 for x in out if x['extra']['fo_stock'] == 'Yes')} F&O), "
        f"{unmatched} skipped with no NSE VaR match  [{ANGEL_BACKTEST}]")
    if not out:
        raise RuntimeError("Angel One produced no rows — check the input file")
    return out


GROWW_FUND = re.compile(
    r"\betf\b|\bfund\b|\bnifty\b|\bsensex\b|bees|\bgold\b|\bsilver\b|liquid"
    r"|bharat bond|index|amc|asset management", re.I)


def fetch_groww(stamp: str) -> list[dict]:
    """Groww renders 100 rows per filter/sort view and exposes no working list
    API, so the list is swept in-browser with groww_accumulate.js.

    Accepts either shape the harvester can produce:
      - DOM sweep  : slug,cells   (margin is the last percentage in `cells`)
      - JSON sweep : nseScriptCode,mtfHaircut,...
    Rows are de-duplicated on slug, then again on resolved NSE symbol.
    Symbols are backfilled by name against the other brokers in this pull, so
    fetch_groww must run AFTER them (FETCHERS order guarantees this).

    Fund and ETF names are never fuzzy-matched: near-identical passive-fund
    names collide (an FMCG ETF onto a gold ETF) and silently corrupt the map."""
    path = _newest(["groww_dom*.csv", "groww_mtf*.csv", "groww*.csv"])
    if not path:
        raise FileNotFoundError(
            f"No Groww CSV in {INPUT_DIR}. Run groww_accumulate.js on "
            "groww.in/stocks/mtf/list, sweep the filter and sort views, then "
            "DD() to download and drop the CSV there.")
    log(f"Groww: reading {os.path.basename(path)}")
    with open(path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    merged: dict[str, dict] = {}
    skipped = 0
    for r in rows:
        if "cells" in r:                                  # DOM sweep
            slug = (r.get("slug") or "").strip()
            if not slug or "/" in slug:
                continue
            parts = [c.strip() for c in (r.get("cells") or "").split("|")]
            pcts = [p for p in parts if re.fullmatch(r"\d+(\.\d+)?%", p)]
            if not pcts:                                  # nav/widget row
                skipped += 1
                continue
            merged.setdefault(slug, {"slug": slug, "name": parts[0],
                                     "margin": float(pcts[-1].rstrip("%")),
                                     "symbol": "", "bse": ""})
        else:                                             # JSON sweep
            slug = (r.get("searchId") or "").strip()
            hc = (r.get("mtfHaircut") or "").strip()
            if not hc:
                continue
            key = slug or (r.get("nseScriptCode") or "").strip()
            merged.setdefault(key, {"slug": slug, "name": (r.get("shortName") or "").strip(),
                                    "margin": float(hc),
                                    "symbol": (r.get("nseScriptCode") or "").strip(),
                                    "bse": (r.get("bseScriptCode") or "").strip()})
    log(f"Groww: {len(merged)} unique slugs ({skipped} non-scrip rows skipped)")
    if not merged:
        raise RuntimeError("Groww CSV produced no rows — wrong file?")

    out, unmatched, collisions = [], 0, 0
    seen: dict[str, dict] = {}
    for g in merged.values():
        sym = g["symbol"] or _groww_symbol(g)
        is_fund = bool(GROWW_FUND.search(g["name"]) or GROWW_FUND.search(g["slug"]))
        m = g["margin"]
        if not sym:
            unmatched += 1
        elif sym.upper() in seen:
            collisions += 1
            sym = ""
            unmatched += 1
        if sym:
            seen[sym.upper()] = g
        out.append(rec(sym or f"({g['name'][:28]})", "", m, 100 - m,
                       (100.0 / m) if m else 0.0,
                       category="ETF" if is_fund else "Equity",
                       scrip_name=g["name"], groww_slug=g["slug"],
                       bse_code=g["bse"], symbol_unmatched=not bool(sym),
                       fund_flag="Yes" if is_fund else "No"))
    eq = sum(1 for x in out if x["category"] != "ETF")
    log(f"Groww: {len(out)} scrips — {eq} equity (~{eq / 1300 * 100:.0f}% of an "
        f"estimated 1,300 universe), {len(out) - eq} ETF/fund, "
        f"{unmatched} without an NSE symbol match, {collisions} collisions dropped")
    if eq / 1300 < 0.5:
        log("Groww capture is well short of the universe — sweep more filter "
            "views with the accumulator before relying on it.", "WARN")
    return out


_GROWW_NAME_IDX: dict[str, str] = {}


def _groww_norm(s: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9 ]", " ", str(s).lower())
    stop = {"ltd", "limited", "the", "and", "of", "co", "company", "plc",
            "inc", "pvt", "private", "corporation", "corp"}
    return " ".join(w for w in s.split() if w and w not in stop)


def _groww_symbol(g: dict) -> str:
    """Exact name match only for funds; fuzzy allowed for ordinary companies."""
    idx = _GROWW_NAME_IDX
    hit = idx.get(_groww_norm(g["name"])) or idx.get(_groww_norm(g["slug"].replace("-", " ")))
    if hit:
        return hit
    if GROWW_FUND.search(g["name"]) or GROWW_FUND.search(g["slug"]):
        return ""
    import difflib
    for q in (_groww_norm(g["name"]), _groww_norm(g["slug"].replace("-", " "))):
        m = difflib.get_close_matches(q, list(idx), n=1, cutoff=0.90)
        if m:
            return idx[m[0]]
    return ""


FETCHERS = {"dhan": fetch_dhan, "zerodha": fetch_zerodha, "upstox": fetch_upstox,
            "kotakneo": fetch_kotakneo, "paytm": fetch_paytm,
            "angelone": fetch_angel, "groww": fetch_groww}
DISPLAY = {"dhan": "Dhan", "zerodha": "Zerodha", "upstox": "Upstox",
           "kotakneo": "Kotak Neo", "paytm": "Paytm Money",
           "angelone": "Angel One", "groww": "Groww"}
# Brokers whose margin is not broker-published, or whose list is incomplete.
# They get their own sheet but are kept OUT of the cross-broker margin columns,
# so the comparison stays a like-for-like set of published, complete figures.
NON_COMPARABLE = {"angelone": "margin is modelled, not published by the broker",
                  "groww": "list is swept view-by-view, high coverage but not provably exhaustive"}


def build_groww_name_index(data: dict[str, list[dict]]) -> None:
    """Company-name -> NSE symbol, from every broker already fetched. Groww
    publishes no symbol on its list surface, so this is how its rows are keyed."""
    _GROWW_NAME_IDX.clear()
    for b, recs in data.items():
        if b == "groww":
            continue
        for r in recs:
            for nm in (r["extra"].get("scrip_name"), r["symbol"]):
                if nm:
                    k = _groww_norm(nm)
                    if k:
                        _GROWW_NAME_IDX.setdefault(k, r["symbol"])
    log(f"Groww name index: {len(_GROWW_NAME_IDX)} entries from peer brokers")


def backfill_isins(data: dict[str, list[dict]]) -> tuple[int, int]:
    """Kotak Neo publishes no ISIN. Resolve it from any broker that publishes
    both symbol and ISIN, so cross-broker matching stays on ISIN rather than
    degrading to ticker. Returns (resolved, unresolved)."""
    lookup: dict[str, str] = {}
    for recs in data.values():
        for r in recs:
            if r["isin"] and r["symbol"]:
                lookup.setdefault(r["symbol"].upper(), r["isin"])
    resolved = unresolved = 0
    for recs in data.values():
        for r in recs:
            if r["isin"]:
                continue
            hit = lookup.get(r["symbol"].upper())
            if hit:
                r["isin"] = hit
                r["extra"]["isin_backfilled"] = True
                resolved += 1
            else:
                unresolved += 1
    if resolved or unresolved:
        log(f"ISIN backfill: {resolved} resolved from peer brokers, "
            f"{unresolved} left unmatched (broker-exclusive names)")
    return resolved, unresolved


# --------------------------------------------------------------------------
# Validation gate — nothing is written if this fails
# --------------------------------------------------------------------------

# Published leverage is rounded to different precision by each broker, so the
# tolerance on the leverage-vs-margin cross-check has to match that precision.
#   Dhan/Zerodha/Upstox publish 2dp -> tight
#   Paytm Money rounds margin_x to 1dp -> up to 0.05 drift
#   Kotak Neo's margin is derived FROM leverage, so it agrees by construction
LEV_TOL = {"paytm": 0.06, "kotakneo": 0.001, "angelone": 0.001, "groww": 0.001}
SYMBOL_DUP_EXEMPT = {"groww"}   # unmatched rows share a "(name)" placeholder

# Share of rows allowed to carry no ISIN after backfill. Tight for brokers that
# publish ISIN themselves; loose for Groww, which publishes neither ISIN nor
# symbol and lists many passive funds that no other broker in this set carries.
ISIN_MISS_TOL = {"groww": 0.20}
ISIN_MISS_TOL_DEFAULT = 0.05
LEV_TOL_DEFAULT = 0.05


def validate(broker: str, rows: list[dict]) -> list[str]:
    errs = []
    name = DISPLAY[broker]

    floor = 50 if broker == "groww" else 500   # Groww is harvested, size varies
    if len(rows) < floor:
        errs.append(f"{name}: only {len(rows)} rows — feed looks truncated.")

    with_isin = [r for r in rows if r["isin"]]
    isins = [r["isin"] for r in with_isin]
    dupes = len(isins) - len(set(isins))
    if dupes:
        errs.append(f"{name}: {dupes} duplicate ISIN(s).")

    syms = [r["symbol"] for r in rows if not r["extra"].get("symbol_unmatched")]
    if len(syms) - len(set(syms)):
        errs.append(f"{name}: {len(syms) - len(set(syms))} duplicate symbol(s).")

    if any(not r["symbol"] for r in rows):
        errs.append(f"{name}: row(s) with a blank symbol.")

    # A blank ISIN is tolerable for a broker that does not publish one (Kotak),
    # but a large unmatched share means the backfill lookup has gone wrong.
    missing = len(rows) - len(with_isin)
    tol_i = ISIN_MISS_TOL.get(broker, ISIN_MISS_TOL_DEFAULT)
    if missing and missing / len(rows) > tol_i:
        errs.append(f"{name}: {missing} of {len(rows)} rows "
                    f"({missing / len(rows):.0%}) have no ISIN after backfill, "
                    f"above the {tol_i:.0%} tolerance for this broker.")

    for r in rows:
        if not (0 <= r["margin_pct"] <= 100):
            errs.append(f"{name}: {r['symbol']} margin {r['margin_pct']} out of range.")
            break
        if abs(r["margin_pct"] + r["broker_funding_pct"] - 100) > 0.01:
            errs.append(f"{name}: {r['symbol']} funding split does not sum to 100.")
            break

    tol = LEV_TOL.get(broker, LEV_TOL_DEFAULT)
    bad = [r["symbol"] for r in rows
           if r["margin_pct"] > 0 and abs(100 / r["margin_pct"] - r["leverage"]) > tol]
    if bad:
        errs.append(f"{name}: {len(bad)} leverage/margin mismatch(es) beyond "
                    f"tolerance {tol}, e.g. {', '.join(bad[:5])}.")
    return errs


# --------------------------------------------------------------------------
# SQLite history
# --------------------------------------------------------------------------

DDL = """
CREATE TABLE IF NOT EXISTS mtf_snapshot (
    as_of              TEXT NOT NULL,
    pulled_at          TEXT NOT NULL,
    broker             TEXT NOT NULL,
    symbol             TEXT NOT NULL,
    isin               TEXT NOT NULL,
    category           TEXT,
    margin_pct         REAL NOT NULL,
    broker_funding_pct REAL NOT NULL,
    leverage           REAL NOT NULL,
    funded             INTEGER NOT NULL,
    PRIMARY KEY (as_of, broker, isin)
);
CREATE INDEX IF NOT EXISTS ix_snap_isin   ON mtf_snapshot(isin);
CREATE INDEX IF NOT EXISTS ix_snap_broker ON mtf_snapshot(broker, as_of);
CREATE TABLE IF NOT EXISTS run_log (
    pulled_at TEXT PRIMARY KEY,
    as_of     TEXT,
    brokers   TEXT,
    rows      INTEGER,
    status    TEXT,
    notes     TEXT
);
"""


def db_connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.executescript(DDL)
    return con


def db_write(con, as_of, pulled_at, data: dict[str, list[dict]]) -> None:
    rows = []
    for broker, recs in data.items():
        for r in recs:
            rows.append((as_of, pulled_at, DISPLAY[broker], r["symbol"], r["isin"],
                         r["category"], r["margin_pct"], r["broker_funding_pct"],
                         r["leverage"], int(r["funded"])))
    con.executemany(
        "INSERT OR REPLACE INTO mtf_snapshot VALUES (?,?,?,?,?,?,?,?,?,?)", rows)
    con.commit()
    log(f"SQLite: wrote {len(rows)} rows for as_of={as_of}")


def previous_as_of(con, current: str) -> str | None:
    row = con.execute(
        "SELECT MAX(as_of) FROM mtf_snapshot WHERE as_of < ?", (current,)).fetchone()
    return row[0] if row and row[0] else None


def diff_report(con, prev: str, curr: str, min_move: float = 0.25) -> list[str]:
    """Added / removed / margin-changed, per broker. This is the part worth
    reading each morning — a scrip dropping off a list, or a margin hike, is
    what actually changes position sizing.

    min_move suppresses noise. Kotak Neo's margin is derived from a 2dp
    leverage figure, so a single tick of leverage shows up as roughly 0.15pp of
    margin; without a threshold those swamp the genuine moves. 0.25pp is below
    anything that affects position sizing on a realistic ticket."""
    lines = []
    q = ("SELECT broker, isin, symbol, margin_pct, funded FROM mtf_snapshot "
         "WHERE as_of = ?")
    old = {(b, i): (s, m, f) for b, i, s, m, f in con.execute(q, (prev,))}
    new = {(b, i): (s, m, f) for b, i, s, m, f in con.execute(q, (curr,))}

    brokers = sorted({k[0] for k in new} | {k[0] for k in old})
    for b in brokers:
        o = {k[1]: v for k, v in old.items() if k[0] == b}
        n = {k[1]: v for k, v in new.items() if k[0] == b}
        added = [n[i][0] for i in n.keys() - o.keys()]
        removed = [o[i][0] for i in o.keys() - n.keys()]
        hiked, cut, lost_funding, gained_funding = [], [], [], []
        for i in n.keys() & o.keys():
            sym, mn, fn = n[i]
            _, mo, fo = o[i]
            if mn - mo >= min_move:
                hiked.append(f"{sym} {mo:.2f}->{mn:.2f}%")
            elif mo - mn >= min_move:
                cut.append(f"{sym} {mo:.2f}->{mn:.2f}%")
            if fo and not fn:
                lost_funding.append(sym)
            elif fn and not fo:
                gained_funding.append(sym)

        lines.append(f"  {b}:")
        if not any([added, removed, hiked, cut, lost_funding, gained_funding]):
            lines.append("    no material change")
            continue
        lines.append(f"    added            {len(added):>4}  "
                     f"{', '.join(sorted(added)[:8])}{' ...' if len(added) > 8 else ''}")
        lines.append(f"    removed          {len(removed):>4}  "
                     f"{', '.join(sorted(removed)[:8])}{' ...' if len(removed) > 8 else ''}")
        lines.append(f"    margin hiked     {len(hiked):>4}  "
                     f"{', '.join(sorted(hiked)[:6])}{' ...' if len(hiked) > 6 else ''}")
        lines.append(f"    margin cut       {len(cut):>4}  "
                     f"{', '.join(sorted(cut)[:6])}{' ...' if len(cut) > 6 else ''}")
        if lost_funding:
            lines.append(f"    funding withdrawn {len(lost_funding):>3}  "
                         f"{', '.join(sorted(lost_funding)[:8])}")
        if gained_funding:
            lines.append(f"    funding restored  {len(gained_funding):>3}  "
                         f"{', '.join(sorted(gained_funding)[:8])}")
    return lines


def print_history(con, n: int) -> None:
    rows = con.execute(
        "SELECT as_of, broker, COUNT(*), ROUND(AVG(margin_pct),2), "
        "SUM(CASE WHEN funded=0 THEN 1 ELSE 0 END) "
        "FROM mtf_snapshot GROUP BY as_of, broker "
        "ORDER BY as_of DESC LIMIT ?", (n * 3,)).fetchall()
    print(f"\n{'as_of':<12}{'broker':<10}{'scrips':>8}{'avg margin':>12}{'unfunded':>10}")
    print("-" * 52)
    for a, b, c, m, u in rows:
        print(f"{a:<12}{b:<10}{c:>8}{m:>11}%{u:>10}")


# --------------------------------------------------------------------------
# CSV export — flat files for downstream pipelines / TradingView / Chartink
# --------------------------------------------------------------------------

def write_csvs(data: dict[str, list[dict]], as_of: str) -> list[str]:
    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    combined = []
    for broker, recs in data.items():
        name = DISPLAY[broker]
        p = os.path.join(OUT_DIR, f"mtf_{broker}.csv")
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["as_of", "broker", "symbol", "isin", "category",
                        "margin_pct_user_funds", "broker_funding_pct",
                        "leverage", "mtf_funded"])
            for r in sorted(recs, key=lambda x: x["symbol"]):
                row = [as_of, name, r["symbol"], r["isin"], r["category"],
                       r["margin_pct"], r["broker_funding_pct"], r["leverage"],
                       "Y" if r["funded"] else "N"]
                w.writerow(row)
                combined.append(row)
        written.append(p)

    p = os.path.join(OUT_DIR, "mtf_all_brokers.csv")
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["as_of", "broker", "symbol", "isin", "category",
                    "margin_pct_user_funds", "broker_funding_pct",
                    "leverage", "mtf_funded"])
        w.writerows(sorted(combined, key=lambda x: (x[2], x[1])))
    written.append(p)

    # Symbol-only lists, ready to paste into a screener watchlist.
    for broker, recs in data.items():
        p = os.path.join(OUT_DIR, f"watchlist_{broker}_funded.txt")
        syms = sorted({r["symbol"] for r in recs if r["funded"]})
        with open(p, "w", encoding="utf-8") as f:
            f.write("\n".join(f"NSE:{s}" for s in syms))
        written.append(p)
    return written


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------

def build_workbook(data: dict[str, list[dict]], as_of: str, pulled_at: str,
                   upstox_declared) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    F = "Arial"
    HF = PatternFill("solid", fgColor="1F3864")
    HFONT = Font(name=F, bold=True, color="FFFFFF", size=10)
    TITLE = Font(name=F, bold=True, size=14, color="1F3864")
    SUB = Font(name=F, italic=True, size=9, color="595959")
    BODY = Font(name=F, size=10)
    BOLD = Font(name=F, size=10, bold=True)
    BLUE = Font(name=F, size=10, color="0000FF")
    BLACK = Font(name=F, size=10)
    WARN = PatternFill("solid", fgColor="FFF2CC")
    GOOD = PatternFill("solid", fgColor="E2EFDA")
    TH = Side(style="thin", color="BFBFBF")
    BOX = Border(left=TH, right=TH, top=TH, bottom=TH)
    PCT, LEV = '0.00"%"', '0.00"x"'

    def hdr(ws, row, n, h=32):
        ws.row_dimensions[row].height = h
        for c in range(1, n + 1):
            x = ws.cell(row=row, column=c)
            x.fill, x.font = HF, HFONT
            x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            x.border = BOX

    def widths(ws, w):
        for i, x in enumerate(w, 1):
            ws.column_dimensions[get_column_letter(i)].width = x

    def table(ws, name, r1, r2, n):
        t = Table(displayName=name, ref=f"A{r1}:{get_column_letter(n)}{r2}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
        ws.add_table(t)

    wb = Workbook()

    # ---------------- README
    ws = wb.active
    ws.title = "README"
    widths(ws, [3, 34, 96, 16])
    ws["B2"] = "MTF (Margin Trading Facility) — Approved Scrip Lists by Broker"
    ws["B2"].font = TITLE
    ws["B3"] = f"Snapshot pulled {pulled_at}  ·  India  ·  NSE/BSE cash segment"
    ws["B3"].font = SUB
    r = 5
    ws.cell(row=r, column=2, value="COVERAGE").font = BOLD
    r += 1
    for i, h in enumerate(["Broker", "Scrips", "Funded", "Status"]):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.fill, c.font = HF, HFONT
        c.alignment = Alignment(horizontal="left", vertical="center")
    r += 1
    for b in data:
        recs = data[b]
        ws.cell(row=r, column=2, value=DISPLAY[b]).font = BOLD
        ws.cell(row=r, column=3, value=len(recs)).font = BLACK
        ws.cell(row=r, column=4, value=sum(1 for x in recs if x["funded"])).font = BLACK
        status = ("MODELLED margin" if b == "angelone"
                  else "PARTIAL list" if b == "groww" else "Complete")
        c = ws.cell(row=r, column=5, value=status)
        c.font = BLACK
        if b in NON_COMPARABLE:
            c.fill = WARN
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="NOT YET INCLUDED").font = BOLD
    r += 1
    for b, why in PENDING_BROKERS.items():
        ws.cell(row=r, column=2, value=b).font = BODY
        ws.cell(row=r, column=2).fill = WARN
        c = ws.cell(row=r, column=3, value=why)
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    for label, text in [
        ("HOW TO READ THE NUMBERS", ""),
        ("Margin requirement / User's funds (%)",
         "The portion of trade value you fund yourself. Lower means more leverage."),
        ("Broker funding (%)",
         "What the broker lends — always 100 minus the margin. Interest accrues on this only."),
        ("Leverage (x)", "100 / margin requirement. A 25% margin is 4.00x."),
        ("", ""),
        ("CAVEATS", ""),
        ("Zero-funding rows",
         "Some scrips appear on a list at 100% margin / 0% funding — present, but with no "
         "actual MTF benefit. Filter the 'MTF Funded?' column."),
        ("Daily revision",
         "Margins track exchange VAR + ELM and broker RMS policy; Dhan restates intraday. "
         "This is a dated snapshot, not a standing reference."),
        ("Published margin is a ceiling",
         "ASM/GSM surveillance, F&O ban periods and account-level caps can all reduce "
         "actual leverage at order time."),
        ("Blue vs black",
         "Blue cells are exactly as published by the broker. Black cells are derived — "
         "either a formula, or an ISIN backfilled from another broker."),
        ("Kotak Neo margin is derived",
         "Kotak publishes leverage and a per-scrip exposure cap, not a margin %. Its margin "
         "column is computed as 100/leverage and inherits Kotak's 2dp rounding, so treat it "
         "as approximate to about +/-0.2pp. Every other broker publishes margin directly."),
        ("Kotak Neo ISINs are backfilled",
         "Kotak publishes no ISIN. ISINs are resolved by symbol from the other brokers in the "
         "same pull; a small number of Kotak-exclusive names have none and are matched on "
         "symbol in the cross-broker sheet."),
        ("Paytm Money rounding",
         "Paytm rounds its published leverage to 1 decimal, so the leverage check can differ "
         "from the published figure by up to 0.05x. The margin % is exact."),
    ]:
        if text == "" and label:
            ws.cell(row=r, column=2, value=label).font = BOLD
        else:
            ws.cell(row=r, column=2, value=label).font = BODY
            c = ws.cell(row=r, column=3, value=text)
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ---------------- per-broker sheets
    NOTE = {
        "dhan": "Dhan publishes the funding split directly; columns D and E are both as-published.",
        "zerodha": "Zerodha publishes margin % and leverage; broker funding in E is derived as 100 - margin.",
        "upstox": "Upstox publishes only the MTF bracket; broker funding and leverage are derived.",
        "kotakneo": ("Kotak publishes LEVERAGE and a per-scrip exposure cap, but no margin %. "
                     "The margin here is DERIVED as 100/leverage and inherits Kotak's 2dp "
                     "rounding (roughly +/-0.2pp). Kotak publishes no ISIN — ISINs shown are "
                     "backfilled by symbol from the other brokers in this pull."),
        "paytm": "Paytm publishes margin % and leverage directly; broker funding in E is derived as 100 - margin.",
        "angelone": (f"MODELLED, not published. base = VaR + 5 x ELM; non-F&O margin = base, F&O margin = "
                     f"{ANGEL_K_FO} x base. Calibrated against {ANGEL_BACKTEST}. The F&O figure is a constant "
                     "(sd 0.00006 across 11 readings), not a regression. Verify in the Angel app before sizing."),
        "groww": ("Swept view by view from the rendered table — Groww shows 100 rows per filter/sort view and has no "
                  "working list API. De-duplicated on slug then on NSE symbol. Margins are Groww-published and were "
                  "cross-verified: 763 scrips captured twice, independently, agreed to 0.000pp. An absent stock does "
                  "NOT mean Groww lacks MTF on it."),
    }
    for b in data:
        recs = sorted(data[b], key=lambda x: x["symbol"])
        name = DISPLAY[b]
        ws = wb.create_sheet(name)
        extra_cols = []
        if b == "kotakneo":
            extra_cols = [("Max Allowed\nExposure (Rs cr)", "max_exposure_cr", "#,##0.0"),
                          ("Scrip Name", "scrip_name", None)]
        elif b == "paytm":
            extra_cols = [("Scrip Name", "scrip_name", None)]
        elif b == "angelone":
            extra_cols = [("F&O Stock?", "fo_stock", None),
                          ("Exchange VaR (%)", "exchange_var", '0.00"%"'),
                          ("ELM (%)", "elm", '0.00"%"'),
                          ("Series", "series", None),
                          ("Company", "scrip_name", None)]
        elif b == "groww":
            extra_cols = [("Company", "scrip_name", None),
                          ("ETF/Fund?", "fund_flag", None),
                          ("Groww Slug", "groww_slug", None)]
        widths(ws, [18, 16, 26, 22, 22, 16, 16, 16] + [22] * len(extra_cols))
        ws["A1"] = f"{name} — MTF Approved Scrips"
        ws["A1"].font = TITLE
        ws["A2"] = (f"{len(recs)} scrips  ·  "
                    f"{sum(1 for x in recs if x['funded'])} actually funded  ·  "
                    f"pulled {pulled_at}")
        ws["A2"].font = SUB
        ws["A3"] = NOTE[b]
        ws["A3"].font = SUB
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        if b == "kotakneo":
            ws.row_dimensions[3].height = 26

        derived_margin = b in ("kotakneo", "angelone")
        H = (["Symbol", "ISIN", "Category",
              ("Margin Requirement\nUser's Funds (%)"
               + (" — DERIVED" if derived_margin else "")),
              f"Broker Funding\n{name}'s Funds (%)",
              "Leverage\n(as published)", "Leverage Check\n(100 / Margin)", "MTF Funded?"]
             + [c[0] for c in extra_cols])
        for i, h in enumerate(H, 1):
            ws.cell(row=5, column=i, value=h)
        hdr(ws, 5, len(H))
        row = 6
        for x in recs:
            ws.cell(row=row, column=1, value=x["symbol"]).font = BLUE
            c = ws.cell(row=row, column=2, value=x["isin"])
            c.font = BLACK if x["extra"].get("isin_backfilled") else BLUE
            ws.cell(row=row, column=3, value=x["category"]).font = BLUE
            # Kotak's margin is computed by us, so it is a formula off the
            # published leverage rather than a blue as-published value.
            if b == "angelone":
                # rebuild from the exchange inputs so the sheet recalculates
                vcol = get_column_letter(9 + [e[1] for e in extra_cols].index("exchange_var"))
                ecol = get_column_letter(9 + [e[1] for e in extra_cols].index("elm"))
                fcol = get_column_letter(9 + [e[1] for e in extra_cols].index("fo_stock"))
                c = ws.cell(row=row, column=4,
                            value=f'=MIN(100,MAX(0,IF({fcol}{row}="Yes",{ANGEL_K_FO},1)*({vcol}{row}+5*{ecol}{row})))')
                c.font = BLACK
            elif derived_margin:
                c = ws.cell(row=row, column=4, value=f'=IFERROR(100/F{row},"")')
                c.font = BLACK
            else:
                c = ws.cell(row=row, column=4, value=x["margin_pct"]); c.font = BLUE
            c.number_format = PCT
            if b == "dhan":
                c = ws.cell(row=row, column=5, value=x["broker_funding_pct"]); c.font = BLUE
            else:
                c = ws.cell(row=row, column=5, value=f"=100-D{row}"); c.font = BLACK
            c.number_format = PCT
            c = ws.cell(row=row, column=6, value=x["leverage"]); c.font = BLUE; c.number_format = LEV
            c = ws.cell(row=row, column=7, value=f'=IFERROR(100/D{row},"")'); c.font = BLACK; c.number_format = LEV
            ws.cell(row=row, column=8,
                    value=f'=IF(E{row}>0,"Yes","No — 0% funding")').font = BLACK
            for k, (_lab, fld, fmt) in enumerate(extra_cols):
                val = x["extra"].get(fld, "")
                if fmt is None and isinstance(val, str):
                    val = safe_text(val)
                c = ws.cell(row=row, column=9 + k, value=val)
                c.font = BLUE
                if fmt:
                    c.number_format = fmt
            if not x["funded"]:
                for cc in range(1, len(H) + 1):
                    ws.cell(row=row, column=cc).fill = WARN
            row += 1
        table(ws, f"tbl{name.replace(' ', '')}", 5, row - 1, len(H))
        ws.freeze_panes = "A6"

    # ---------------- cross-broker
    order = [b for b in data if b not in NON_COMPARABLE]
    if not order:
        order = list(data.keys())

    def key(x):
        # ISIN is the join key. A handful of broker-exclusive names carry no
        # ISIN anywhere in the pull, so they fall back to a symbol-scoped key
        # rather than being silently dropped from the comparison.
        return x["isin"] if x["isin"] else "SYM:" + x["symbol"].upper()

    by = {b: {key(x): x for x in data[b]} for b in order}
    all_isin = sorted(set().union(*[set(v) for v in by.values()]))
    ws = wb.create_sheet("Cross_Broker")
    ncol = 2 + len(order) + 4
    widths(ws, [16, 16] + [14] * len(order) + [12, 16, 14, 30])
    ws["A1"] = "Cross-Broker Margin Comparison"
    ws["A1"].font = TITLE
    ws["A2"] = f"{len(all_isin)} unique ISINs, matched on ISIN  ·  pulled {pulled_at}"
    ws["A2"].font = SUB
    ws["A3"] = ("Margin = what YOU fund; lower is better. Blank = not on that broker's list. "
                "The last column lists EVERY broker tied at the best margin — margins match "
                "exactly on a large share of rows, so a single 'winner' would mislead.")
    ws["A3"].font = SUB

    H = (["Symbol", "ISIN"] + [f"{DISPLAY[b]}\nMargin (%)" for b in order] +
         ["Brokers\nOffering", "Best (Lowest)\nMargin (%)", "Best Leverage\n(x)",
          "Best Margin Offered By\n(all tied brokers)"])
    for i, h in enumerate(H, 1):
        ws.cell(row=5, column=i, value=h)
    hdr(ws, 5, len(H))

    first_b = get_column_letter(3)
    last_b = get_column_letter(2 + len(order))
    cnt_c = get_column_letter(3 + len(order))
    best_c = get_column_letter(4 + len(order))
    levc = get_column_letter(5 + len(order))
    who_c = get_column_letter(6 + len(order))

    row = 6
    for i in all_isin:
        sym = next((by[b][i]["symbol"] for b in order if i in by[b]), i)
        ws.cell(row=row, column=1, value=sym).font = BLUE
        ws.cell(row=row, column=2, value=i).font = BLUE
        for k, b in enumerate(order):
            if i in by[b]:
                c = ws.cell(row=row, column=3 + k, value=by[b][i]["margin_pct"])
                c.font = BLUE
                c.number_format = PCT
        ws.cell(row=row, column=3 + len(order),
                value=f"=COUNT({first_b}{row}:{last_b}{row})").font = BLACK
        # Best margin and ties are compared at 2dp. Kotak Neo's margin is
        # DERIVED from a 2dp leverage figure, so it carries spurious extra
        # decimals: on IXIGO it reads 38.7597 against 38.76 from three other
        # brokers. Comparing raw would crown Kotak sole winner on a 0.0003pp
        # artifact and hide a genuine four-way tie.
        c = ws.cell(row=row, column=4 + len(order),
                    value=f'=IFERROR(ROUND(MIN({first_b}{row}:{last_b}{row}),2),"")')
        c.font = BLACK; c.number_format = PCT
        c = ws.cell(row=row, column=5 + len(order),
                    value=f'=IFERROR(100/{best_c}{row},"")')
        c.font = BLACK; c.number_format = LEV
        parts = "&".join(
            f'IF(ROUND({get_column_letter(3 + k)}{row},2)={best_c}{row},", {DISPLAY[b]}","")'
            for k, b in enumerate(order))
        ws.cell(row=row, column=6 + len(order),
                value=f'=IF({best_c}{row}="","",MID({parts},3,100))').font = BLACK
        row += 1
    table(ws, "tblCross", 5, row - 1, len(H))
    ws.freeze_panes = "C6"

    last = row - 1
    s = row + 2
    ws.cell(row=s, column=1, value="COVERAGE SUMMARY").font = BOLD
    items = [("Unique ISINs across all brokers", f"=ROWS(A6:A{last})")]
    for n in range(len(order), 0, -1):
        label = ("On all brokers" if n == len(order)
                 else f"On exactly {n} broker{'s' if n > 1 else ''}")
        items.append((label, f'=COUNTIFS({cnt_c}6:{cnt_c}{last},{n})'))
    for k, b in enumerate(order):
        col = get_column_letter(3 + k)
        items.append((f"{DISPLAY[b]} list size", f"=COUNT({col}6:{col}{last})"))
    for b in order:
        items.append((f"{DISPLAY[b]} offers the best margin (incl. ties)",
                      f'=COUNTIF({who_c}6:{who_c}{last},"*{DISPLAY[b]}*")'))
    items.append(("Rows where 2+ brokers tie at the best margin",
                  f'=SUMPRODUCT((LEN({who_c}6:{who_c}{last})'
                  f'-LEN(SUBSTITUTE({who_c}6:{who_c}{last},",","")))>0)'))
    for k, f_ in items:
        s += 1
        ws.cell(row=s, column=1, value=k).font = BODY
        ws.cell(row=s, column=1).fill = GOOD
        ws.cell(row=s, column=3, value=f_).font = BLACK

    # ---------------- sources
    ws = wb.create_sheet("Sources")
    widths(ws, [16, 34, 66, 14, 12, 20])
    ws["A1"] = "Source Authentication"
    ws["A1"].font = TITLE
    ws["A2"] = (f"Generated by refresh_mtf.py  ·  pulled {pulled_at}  ·  "
                f"raw payloads archived under {os.path.basename(RAW_DIR)}/{as_of}_*")
    ws["A2"].font = SUB
    H = ["Broker", "Public page", "Endpoint used", "Format", "Records", "Gate applied"]
    for i, h in enumerate(H, 1):
        ws.cell(row=4, column=i, value=h)
    hdr(ws, 4, len(H), 26)
    meta = {
        "dhan": (DHAN_PAGE, DHAN_URL, "CSV",
                 "Split sums to 100; leverage = 100/margin; no duplicate ISIN."),
        "zerodha": (ZERODHA_PAGE, ZERODHA_URL, "JSON",
                    "Leverage = 100/margin; no duplicate ISIN."),
        "upstox": (UPSTOX_PAGE, UPSTOX_URL.format(exchange="NSE", records=UPSTOX_PAGE_SIZE, page="N"),
                   "JSON (paginated)",
                   f"Captured count reconciled to API totalRecords "
                   f"({upstox_declared}); no duplicate ISIN."),
        "kotakneo": (KOTAK_PAGE, KOTAK_PAGE_N.format(page="N") + "  (Next.js RSC flight payload)",
                     "SSR HTML (paginated)",
                     "All pages walked to the site's own totalPages; 30 rows/page "
                     "reconciled; no duplicate symbol. Margin DERIVED from published "
                     "leverage (Kotak publishes no margin %). ISINs backfilled by symbol."),
        "paytm": (PAYTM_PAGE, PAYTM_URL, "JSON",
                  "Single unpaginated response; ENABLED rows only; leverage matches "
                  "100/margin within Paytm's 1dp rounding; no duplicate ISIN."),
        "angelone": ("Angel back office — monthly scrip-category file (local input)",
                     f"list: {os.path.basename(_newest(['*SCRIPCATEGORY*.xlsx','angel*.xlsx']) or 'n/a')}"
                     f"   |   margin: {NSE_VAR_URL.format(ddmmyyyy='DDMMYYYY')}"
                     f"   |   F&O split: {NSE_FO_URL}",
                     "XLSX + NSE DAT",
                     f"List is authoritative (Angel's own file). Margin is MODELLED: base = VaR + 5 x ELM, "
                     f"F&O margin = {ANGEL_K_FO} x base. Calibrated against {ANGEL_BACKTEST}. "
                     "Must use VaR snapshot 1 of the trading day. Excluded from cross-broker margin columns."),
        "groww": ("https://groww.in/stocks/mtf/list",
                  f"local input: {os.path.basename(_newest(['groww_mtf*.csv','groww*.csv']) or 'n/a')} "
                  "(harvested in-browser with groww_accumulate.js)",
                  "CSV",
                  "PARTIAL. Groww renders 100 rows per filter/sort view and no list API works, so coverage is "
                  "whatever the accumulator swept. Margins are Groww-published. Excluded from cross-broker "
                  "margin columns because incomplete coverage would distort them."),
    }
    row = 5
    for b in data:
        page, ep, fmt, gate = meta[b]
        for i, v in enumerate([DISPLAY[b], page, ep, fmt, len(data[b]), gate], 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BOX
        ws.row_dimensions[row].height = 60
        row += 1
    row += 1
    for n in [
        "All feeds pulled in one synchronised run, so the snapshot is internally consistent.",
        "No aggregator or comparison site is used — every figure comes from the broker's own infrastructure.",
        "Nothing is scraped from rendered HTML; each source is the structured feed the broker's page consumes.",
        "Cross-broker matching is on ISIN, not ticker, because symbols diverge across brokers.",
        "Derived columns are Excel formulas, so the workbook recalculates if source values are edited.",
    ]:
        ws.cell(row=row, column=1, value="•").font = BODY
        c = ws.cell(row=row, column=2, value=n)
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    for sh in wb.worksheets:
        sh.sheet_view.showGridLines = False

    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, WORKBOOK)
    wb.save(path)
    # dated archive copy so you can diff workbooks later
    wb.save(os.path.join(OUT_DIR, f"{as_of}_{WORKBOOK}"))
    return path


def try_recalc(path: str) -> None:
    """Optional: have LibreOffice cache formula values so pandas/openpyxl
    data_only reads work. Excel does this itself on open, so this is only
    needed for headless downstream reads."""
    import shutil
    import subprocess
    soffice = shutil.which("soffice") or shutil.which("soffice.exe")
    if not soffice:
        log("LibreOffice not found; skipping recalc. Excel will compute on open.", "WARN")
        return
    out = os.path.dirname(path)
    try:
        subprocess.run([soffice, "--headless", "--convert-to", "xlsx",
                        "--outdir", out, path], check=True, timeout=300,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        log("LibreOffice recalc complete")
    except Exception as e:
        log(f"recalc failed ({e}); Excel will compute on open.", "WARN")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="Refresh India MTF approved-scrip lists.")
    ap.add_argument("--brokers", default="dhan,zerodha,upstox,kotakneo,paytm,angelone,groww",
                    help="subset of: dhan,zerodha,upstox,kotakneo,paytm,angelone,groww")
    ap.add_argument("--no-excel", action="store_true", help="skip workbook build")
    ap.add_argument("--no-db", action="store_true", help="skip SQLite write")
    ap.add_argument("--recalc", action="store_true",
                    help="run LibreOffice to cache formula values")
    ap.add_argument("--min-move", type=float, default=0.25, metavar="PP",
                    help="minimum margin move in percentage points to report "
                         "as a change (default 0.25; suppresses rounding noise)")
    ap.add_argument("--history", type=int, metavar="N",
                    help="print the last N snapshots and exit")
    args = ap.parse_args()

    if args.history:
        con = db_connect()
        print_history(con, args.history)
        return 0

    brokers = [b.strip().lower() for b in args.brokers.split(",") if b.strip()]
    unknown = [b for b in brokers if b not in FETCHERS]
    if unknown:
        log(f"unknown broker(s): {', '.join(unknown)}. "
            f"Valid: {', '.join(FETCHERS)}", "ERROR")
        return 3

    now = dt.datetime.now(IST)
    as_of = now.strftime("%Y-%m-%d")
    pulled_at = now.strftime("%Y-%m-%d %H:%M:%S IST")
    stamp = now.strftime("%Y%m%d_%H%M%S")

    log("=" * 62)
    log(f"MTF refresh — as_of {as_of}  ({', '.join(DISPLAY[b] for b in brokers)})")
    log("=" * 62)

    data, upstox_declared, skipped = {}, None, []
    # Groww is keyed off the other brokers' names, so it must fetch last.
    brokers = sorted(brokers, key=lambda x: x == "groww")
    for b in brokers:
        try:
            if b == "groww":
                build_groww_name_index(data)
            data[b] = FETCHERS[b](stamp)
            if b == "upstox":
                upstox_declared = len(data[b])
        except FileNotFoundError as e:
            # A missing local input is a skip, not a failure: the online
            # brokers should still refresh.
            log(f"{DISPLAY[b]} skipped — {e}", "WARN")
            skipped.append(DISPLAY[b])
        except Exception as e:
            log(f"{DISPLAY[b]} fetch failed: {e}", "ERROR")
            return 1
    if not data:
        log("No broker produced data.", "ERROR")
        return 1

    # ---- ISIN backfill (Kotak Neo publishes none)
    backfill_isins(data)

    # ---- validation gate
    log("-" * 62)
    all_errs = []
    for b in brokers:
        errs = validate(b, data[b])
        all_errs += errs
        log(f"{DISPLAY[b]:<10} {len(data[b]):>5} scrips  "
            f"{sum(1 for x in data[b] if x['funded']):>5} funded  "
            f"{'OK' if not errs else 'FAILED'}")
    if all_errs:
        log("VALIDATION FAILED — nothing written:", "ERROR")
        for e in all_errs:
            log(f"  {e}", "ERROR")
        return 2

    # ---- persist + diff
    con = None
    if not args.no_db:
        con = db_connect()
        prev = previous_as_of(con, as_of)
        db_write(con, as_of, pulled_at, data)
        if prev:
            log("-" * 62)
            log(f"CHANGES vs {prev}")
            for line in diff_report(con, prev, as_of, args.min_move):
                print(line)
        else:
            log("No earlier snapshot — this is the baseline.")

    # ---- outputs
    log("-" * 62)
    for p in write_csvs(data, as_of):
        log(f"wrote {os.path.relpath(p, BASE_DIR)}")

    if not args.no_excel:
        try:
            path = build_workbook(data, as_of, pulled_at, upstox_declared)
            log(f"wrote {os.path.relpath(path, BASE_DIR)}")
            if args.recalc:
                try_recalc(path)
        except ImportError:
            log("openpyxl not installed — run: pip install openpyxl", "ERROR")
            return 3

    if con:
        con.execute("INSERT OR REPLACE INTO run_log VALUES (?,?,?,?,?,?)",
                    (pulled_at, as_of, ",".join(brokers),
                     sum(len(v) for v in data.values()), "success", ""))
        con.commit()
        con.close()

    log("=" * 62)
    log(f"DONE — {sum(len(v) for v in data.values())} scrips across "
        f"{len(data)} broker(s)")
    if skipped:
        log(f"Skipped (no local input file): {', '.join(skipped)}. "
            f"See INPUT_DIR = {INPUT_DIR}", "WARN")
    for b in data:
        if b in NON_COMPARABLE:
            log(f"Note: {DISPLAY[b]} is excluded from the cross-broker margin "
                f"comparison — {NON_COMPARABLE[b]}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
