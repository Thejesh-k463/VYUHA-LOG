"""Independent audit of the finished workbook."""
from openpyxl import load_workbook
import re,json,sys,statistics as st
from decimal import Decimal,ROUND_HALF_UP
def xlround(v,n=2):
    # Excel rounds half AWAY FROM ZERO; Python's round() is banker's rounding.
    return float(Decimal(repr(v)).quantize(Decimal('1.'+'0'*n),rounding=ROUND_HALF_UP))
WB=sys.argv[1]
wb=load_workbook(WB,data_only=True)
BROKERS=['Dhan','Zerodha','Upstox','Kotak Neo','Paytm Money']
def rng(ws):
    t=list(ws.tables.values())[0]
    return 6,int(re.sub(r'[A-Z]','',t.ref.split(':')[1]))
fails=[]
print("SHEETS:",wb.sheetnames)
print("\n1. ROW COUNTS")
for sh,e in [('Dhan',1742),('Zerodha',1493),('Upstox',1440),('Kotak Neo',1680),
             ('Paytm Money',1460),('Angel One',1557),('Groww',1292)]:
    lo,hi=rng(wb[sh]); n=hi-lo+1
    ok=n==e; print(f"   {sh:<13}{n:>6} expected {e:>6} {'OK' if ok else 'FAIL'}")
    if not ok: fails.append(f"{sh} rows")
print("\n2. DUPLICATES (within table)")
for sh in ['Dhan','Zerodha','Upstox','Kotak Neo','Paytm Money','Angel One','Groww']:
    ws=wb[sh]; lo,hi=rng(ws); sy=[];isn=[]
    for r in range(lo,hi+1):
        a=ws.cell(r,1).value; b=ws.cell(r,2).value
        if isinstance(a,str) and not a.startswith('('): sy.append(a)
        if isinstance(b,str) and b.startswith('INE'): isn.append(b)
    ds=len(sy)-len(set(sy)); di=len(isn)-len(set(isn))
    print(f"   {sh:<13} dupSym {ds:>3}  dupISIN {di:>3} {'OK' if not(ds or di) else 'FAIL'}")
    if ds or di: fails.append(f"{sh} dup")
print("\n3. MARGIN RANGE + FUNDING SPLIT")
for sh,mc,fc in [('Dhan',4,5),('Zerodha',4,5),('Upstox',4,5),('Kotak Neo',4,5),
                 ('Paytm Money',4,5),('Angel One',4,5),('Groww',4,5)]:
    ws=wb[sh]; lo,hi=rng(ws); bad=0;n=0
    for r in range(lo,hi+1):
        m=ws.cell(r,mc).value
        if not isinstance(m,(int,float)): continue
        n+=1
        if not (0<m<=100): bad+=1; continue
        if fc:
            f=ws.cell(r,fc).value
            if isinstance(f,(int,float)) and abs(m+f-100)>0.02: bad+=1
    print(f"   {sh:<13} n {n:>5}  bad {bad:>3} {'OK' if bad==0 else 'FAIL'}")
    if bad: fails.append(f"{sh} margin")
print("\n4. CROSS-BROKER")
ws=wb['Cross_Broker']; hdr=[str(ws.cell(5,c).value).replace('\n',' ') for c in range(1,ws.max_column+1)]
mc=[i for i,h in enumerate(hdr,1) if any(h.startswith(b+' Margin') for b in BROKERS)]
cc=hdr.index('Brokers Offering')+1; bc=[i for i,h in enumerate(hdr,1) if h.startswith('Best (Lowest)')][0]
lc=[i for i,h in enumerate(hdr,1) if h.startswith('Best Leverage')][0]
wc=[i for i,h in enumerate(hdr,1) if h.startswith('Best Margin Offered By')][0]
lo,hi=rng(ws); b1=b2=b3=b4=0;rows=0
for r in range(lo,hi+1):
    nn=[ws.cell(r,c).value for c in mc]; nn=[v for v in nn if isinstance(v,(int,float))]
    if not nn: continue
    rows+=1
    if ws.cell(r,cc).value!=len(nn): b1+=1
    b=ws.cell(r,bc).value
    if isinstance(b,(int,float)):
        if abs(b-xlround(min(nn)))>0.001: b2+=1
        l=ws.cell(r,lc).value
        if isinstance(l,(int,float)) and abs(l-100/b)>0.01: b3+=1
        tied={BROKERS[mc.index(c)] for c in mc if isinstance(ws.cell(r,c).value,(int,float)) and abs(xlround(ws.cell(r,c).value)-b)<0.001}
        got={x.strip() for x in str(ws.cell(r,wc).value or '').split(',') if x.strip()}
        if tied!=got: b4+=1
print(f"   rows {rows}  count {b1}  best {b2}  leverage {b3}  ties {b4}")
if b1 or b2 or b3 or b4: fails.append("cross-broker")
mg=[h for h in hdr if ('Angel' in h or 'Groww' in h) and 'Margin' in h]
print(f"   modelled/partial brokers in margin cols: {mg or 'none — correct'}")
if mg: fails.append("cross-broker contamination")
print("\n5. ANGEL MODEL BACK-TEST")
obs=json.load(open('raw/angel_allobs.json')); ws=wb['Angel One']; lo,hi=rng(ws); e=[]
for r in range(lo,hi+1):
    s0=ws.cell(r,1).value
    if s0 in obs and isinstance(ws.cell(r,4).value,(int,float)): e.append(abs(ws.cell(r,4).value-obs[s0]))
print(f"   n {len(e)}  MAE {sum(e)/len(e):.4f}pp  max {max(e):.4f}pp  {'OK' if max(e)<0.1 else 'FAIL'}")
if max(e)>=0.1: fails.append("angel model")
print("\n6. GROWW vs ZERODHA (independent)")
zer={x['tradingsymbol'].upper():x['margin'] for x in json.load(open('raw/zerodha.json'))}
ws=wb['Groww']; lo,hi=rng(ws); d=[]
for r in range(lo,hi+1):
    s0=ws.cell(r,1).value; m=ws.cell(r,4).value
    if isinstance(s0,str) and isinstance(m,(int,float)) and s0.upper() in zer: d.append(m-zer[s0.upper()])
print(f"   overlap {len(d)}  median {st.median(d):+.2f}pp  within 1pp {sum(1 for x in d if abs(x)<=1)/len(d):.0%}")
print("\n"+("="*50)+f"\nAUDIT: {'PASS — no issues' if not fails else 'FAIL: '+', '.join(fails)}")
